import impyute
import numpy as np
import random
import openpyxl
from openpyxl.styles.fills import PatternFill
import math
from sklearn.experimental import enable_iterative_imputer
from sklearn.impute import IterativeImputer, KNNImputer
import json
from matplotlib import pyplot as plt


def read_csv(filename : str) -> np.array:
    """
    Returns a Numpy array of the evaluated data

    input:
        filename: string of the path to the file to read

    output:
        names: names of each column
        np.array: read data
    """
    with open(filename, 'r') as f:
        lines = f.read().split()
        f.close()
    names   = [eval(cpg_name) for cpg_name in lines[0].split(',')]
    data    = [line.split(',')[1:] for line in lines[1:]]
    for l_index, line in enumerate(data):
        for v_index, value in enumerate(line):
            try:
                data[l_index][v_index] = float(value)
            except ValueError:
                data[l_index][v_index] = np.nan
    return names, np.array(data)

def get_non_nan(array : np.ndarray, start_n : int = 0, col_n : int = 5) -> np.array:
    cols = array.transpose()
    cols_nan = np.isnan(cols)
    n_cols_nan = [np.count_nonzero(col) for col in cols_nan]
    m = sorted(n_cols_nan)[start_n:start_n+col_n]
    col_indicies = []
    row_indicies = [i for i in range(array.shape[0])]
    new_cols = []
    for i, value in enumerate(n_cols_nan):
            if value in m:
                col_indicies.append(i)
                new_cols.append(cols[i])
    new_rows = np.array(new_cols).transpose()
    new_array = new_rows.copy()
    deleted = 0
    for i, row in enumerate(np.isnan(new_rows)):
        if np.count_nonzero(row) > 0:
            new_array = np.delete(new_array, i-deleted, axis=0)
            row_indicies.remove(i)
            deleted += 1
    return row_indicies, col_indicies, new_array

def get_small_nan(array : np.ndarray, min_n : int = 5, square : bool = True, use_mult : bool =False) -> np.array:
    height, width = array.shape
    if use_mult:
        if square:
            mult_height = []
            mult_width = []
        n_height, n_width = min_n, min_n
        for x in range(n_height, height+1):
            if height%x==0:
                if square:
                    mult_height.append(x)
                else:
                    n_height = x
                    break
        for x in range(n_width, width+1):
            if width%x==0:
                if square:
                    mult_width.append(x)
                else:
                    n_width = x
                    break
        if square:
            common_mult = set(mult_width).intersection(set(mult_height))
            if len(common_mult) != 0:
                n_width, n_height = list(common_mult)[0], list(common_mult)[0]
            else:
                n_width, n_height = mult_width[0], mult_height[0]
    else:
        n_width, n_height = min_n, min_n
    isnan       = np.isnan(array)
    segments    = []
    n_segments  = []
    rows_cols   = []
    for i in range(1, height//n_height+1):
        for j in range(1, width//n_width+1):
            nan_arr = isnan[((i-1)*n_height+1):(i*n_height+1), ((j-1)*n_width+1):(j*n_width+1)]
            arr = array[((i-1)*n_height+1):(i*n_height+1), ((j-1)*n_width+1):(j*n_width+1)]
            rows_cols.append(((((i-1)*n_height+1), ((j-1)*n_width+1)), ((i*n_height+1), (j*n_width+1))))
            segments.append(arr)
            n_segments.append(np.count_nonzero(nan_arr))
    m = min(n_segments)
    indexes = []
    for index, n in enumerate(n_segments):
        if n == m:
            indexes.append(index)
    return len(indexes), [rows_cols[index] for index in indexes], [segments[index] for index in indexes]

def save_to_sheet(array : np.ndarray, sheet, diffs : dict, names : list = None, samples : list = [], highlight : list = []):
    """
    Save numpy array to an excel sheet with specified highlighted nan coordinates
    """
    # If names of columns are provided add them to the first row
    if names != None:
        names = ['Sample #']+names
        sheet.append(names)
    # For each row replace nan with 'NA' string
    for n, row in enumerate(array):
        l_row = list(row)
        for i, value in enumerate(np.isnan(row)):
            if value:
                l_row[i] = 'NA'
        row = [f'Sample #{samples[n]}']+l_row
        sheet.append(row)
    for cell in highlight:
        sheet[chr(65+cell[0]+1)+str(cell[1]+2)].fill = PatternFill(start_color='ebe834', end_color='ebe834', fill_type='solid')
    sheet.append([])
    sheet.append([f'Percentage diff {chr(65+highlight[0][0]+1)+str(highlight[0][1]+2)}', diffs['Percentage diffs'][0], '', 'Sum of percentage diffs', diffs["Percentage diffs sum"], 'Average percentage diff', diffs["Average percentage diff"], 'RMSE', diffs['RMSE']])
    for i, cell in enumerate(highlight[1:]):
        sheet.append([f'Percentage diff {chr(65+cell[0]+1)+str(cell[1]+2)}', diffs["Percentage diffs"][i+1]])
    return sheet
    

def randints(max : int, n_rand : int, seed : int = 25) -> list:
    nums = [None for _ in range(n_rand)]
    for i in range(n_rand):
        random.seed(seed+i)
        nums[i] = round(random.random()*max)-1
    return nums

def get_vals(array : np.ndarray, indicies : list) -> list:
    return [array[i[1], i[0]] for i in indicies]

def get_diffs(og_vals : list, imp: np.ndarray, indicides: list):
    n_nan = len(indicides)
    new_vals = get_vals(imp, indicides)
    perc_diff = []
    sum_diff = 0
    av_diff = 0
    for i in range(n_nan):
        diff = 0
        if og_vals[i] != 0:
            diff = abs(((new_vals[i]-og_vals[i])/og_vals[i])*100)
        perc_diff.append(diff)
        sum_diff += diff
    av_diff = sum_diff/n_nan
    return {'Percentage diffs': perc_diff, 'Percentage diffs sum': sum_diff, 'Average percentage diff': av_diff}

def rmse(og_vals : list, imp: np.ndarray, indicides: list):
    n_nan = len(indicides)
    new_vals = get_vals(imp, indicides)
    sse = 0
    for i, val in enumerate(new_vals):
        sse += (val-og_vals[i])**2
    mse = sse/n_nan
    return math.sqrt(mse)

def test_all_diff(array : np.ndarray, imputation_funcs : dict, n_nan : int = 1, names : list = None, samples : list = [], filename : str = 'Test', seed : int = 25):
    height, width = array.shape
    col_indicies = randints(width, n_nan, seed)
    row_indicies = randints(height, n_nan, seed+n_nan)
    na_indicies = [(col_indicies[i], row_indicies[i]) for i in range(n_nan)]
    og_vals = get_vals(array, na_indicies)
    for index in na_indicies:
        array[index[1], index[0]] = np.nan

    res = {}

    for method, func in imputation_funcs.items():
        imputed = func(array)
        diffs = get_diffs(og_vals,imputed, na_indicies)
        diffs['RMSE'] = rmse(og_vals, imputed, na_indicies)
        res[method] = diffs
    
    return res

def test_all_excel(array : np.ndarray, imputation_funcs : dict, n_nan : int = 1, names : list = None, samples : list = [], filename : str = 'Test', seed : int = 25) -> list:
    """
    Test all relevant imputation methods and save them to a spreadsheet
    """
    height, width = array.shape
    col_indicies = randints(width, n_nan, seed)
    row_indicies = randints(height, n_nan, seed+n_nan)
    na_indicies = [(col_indicies[i], row_indicies[i]) for i in range(n_nan)]
    og_vals = get_vals(array, na_indicies)
    # Create workbook to store data
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Original'
    sheet = save_to_sheet(array, sheet, ([0 for _ in range(n_nan)],0,0,0), names, samples, na_indicies)
    for index in na_indicies:
        array[index[1], index[0]] = np.nan

    for method, func in imputation_funcs.items():
        workbook.create_sheet(title=method)
        sheet = workbook[method]
        imputed = func(array)
        diffs = get_diffs(og_vals,imputed, na_indicies)
        diffs['RMSE'] = rmse(og_vals, imputed, na_indicies)
        sheet = save_to_sheet(imputed, sheet, diffs, names, samples, na_indicies)

    workbook.save(f'{filename}.xlsx')


if __name__ == '__main__':
    from pprint import pformat
    np.set_printoptions(linewidth=1000)
    names, array = read_csv('sample_data_raf.csv')
    # smallest_nan = get_small_nan(array, min_n=12)

    it_imputer_25_0 = IterativeImputer(max_iter=25, random_state=0, tol=0.001)
    it_imputer_25_1 = IterativeImputer(max_iter=25, random_state=1)
    it_imputer_100_1 = IterativeImputer(max_iter=100, random_state=1)
    it_imputer_100_0 = IterativeImputer(max_iter=100, random_state=0)
    knn_imputer_2 = KNNImputer(n_neighbors=2)
    knn_imputer_4 = KNNImputer(n_neighbors=4)
    knn_imputer_6 = KNNImputer(n_neighbors=6)
    knn_imputer_8 = KNNImputer(n_neighbors=8)
    knn_imputer_10 = KNNImputer(n_neighbors=10)
    knn_imputer_12 = KNNImputer(n_neighbors=12)
    knn_imputer_14 = KNNImputer(n_neighbors=14)

    final_all = {}
    final_rmse = {}
    total_rmse = {}
    rmse_prog = {}

    start = 2
    step = 2
    n_steps = 100

    imputations = {
            # 'MICE impyute': impyute.mice,
            # 'EM impyute': impyute.em,
            'Iterative scilearn max=25 random=0': it_imputer_25_0.fit_transform,
            'Iterative scilearn max=25 random=1': it_imputer_25_1.fit_transform,
            # 'Iterative scilearn max=100 random=0': it_imputer_100_0.fit_transform,
            # 'Iterative scilearn max=100 random=1': it_imputer_100_1.fit_transform,
            # 'KNN scilearn k=2': knn_imputer_2.fit_transform,
            # 'KNN scilearn k=4': knn_imputer_4.fit_transform,
            # 'KNN scilearn k=6': knn_imputer_6.fit_transform,
            # 'KNN scilearn k=8': knn_imputer_8.fit_transform,
            # 'KNN scilearn k=10': knn_imputer_10.fit_transform,
            # 'KNN scilearn k=12': knn_imputer_12.fit_transform,
            # 'KNN scilearn k=14': knn_imputer_14.fit_transform
        }

    for i in range(start,start+(n_steps-1)*step+1,step):
        no_nan = get_non_nan(array, 0, i)
        if no_nan[2].shape[0]*no_nan[2].shape[1] == 0:
            n_steps -= n_steps-(i/step-1)
            break
        sample_size = no_nan[2].shape[0]*no_nan[2].shape[1]
        res = test_all_diff(no_nan[2], imputations, 2, [names[i] for i in no_nan[1]], no_nan[0],seed=25)
        final_all[f'{i} columns {sample_size}'] = res
        final_rmse[f'{i} columns {sample_size}'] = {key: value['RMSE'] for key, value in res.items()}
        for key, value in res.items():
            if key in total_rmse:
                total_rmse[key] += value['RMSE']
            else:
                total_rmse[key] = value['RMSE']
            if key not in rmse_prog:
                rmse_prog[key] = {}
            rmse_prog[key][i] = value['RMSE']

    final_rmse['Average RMSE'] = {}

    for key, value in total_rmse.items():
        final_rmse['Average RMSE'][key] = value/n_steps

    with open('result_all.json', 'w') as f:
        f.write(json.dumps(final_all, indent=4))
        f.close()
    with open('result_rmse.json', 'w') as f:
        f.write(json.dumps(final_rmse, indent=4))
        f.close()
    with open('result_rmse_prog.json', 'w') as f:
        f.write(json.dumps(rmse_prog, indent=4))
        f.close()
    
    plot = 'Iterative scilearn max=25 random=0'
    plt.title(plot)
    plt.plot([x for x in rmse_prog[plot]],[y for _, y in rmse_prog[plot].items()])  

    # fig, axs = plt.subplots(3, 4)
    # for i, (key, value) in enumerate(rmse_prog.items()):
    #     axs[(i//4), i-(i//4)*4].set_title(key)
    #     axs[(i//4), i-(i//4)*4].plot([x for x in value], [value[x] for x in value])   
    plt.show()